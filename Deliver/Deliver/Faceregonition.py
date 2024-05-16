import tkinter as tk
from tkinter import Message ,Text
import cv2,os
import shutil
import csv
import numpy as np
from PIL import Image, ImageTk
import pandas as pd
import datetime
import time
import tkinter.ttk as ttk
import tkinter.font as font
import sqlite3
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter,column_index_from_string

url = "https://www.fast2sms.com/dev/bulk"

recognizer = cv2.face.LBPHFaceRecognizer_create()

cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);
from_date = datetime.datetime.today()
currentDate = time.strftime("%d_%m_%y")
font = cv2.FONT_HERSHEY_SIMPLEX
fontScale=1
fontColor=(255,255,255)
cond=0


window = tk.Tk()
window.title("Face_Recogniser")

 
window.geometry('1280x720')
window.configure(background='blue')
#window.attributes('-fullscreen', True)

window.grid_rowconfigure(0, weight=1)
window.grid_columnconfigure(0, weight=1)


message1 = tk.Label(window, text="Face Recognition and Attendance" ,bg="blue"  ,fg="white"  ,width=50  ,height=3,font=('times', 30, 'italic bold underline')) 
message1.place(x=100, y=20)

lbl = tk.Label(window, text="Enter ID",width=20  ,height=2  ,fg="red"  ,bg="yellow" ,font=('times', 15, ' bold ') ) 
lbl.place(x=100, y=200)

txt = tk.Entry(window,width=20,bg="yellow" ,fg="red",font=('times', 15, ' bold '))
txt.place(x=400, y=215)

lbl2 = tk.Label(window, text="Enter Name",width=20  ,fg="red"  ,bg="yellow"    ,height=2 ,font=('times', 15, ' bold ')) 
lbl2.place(x=100, y=300)

txt2 = tk.Entry(window,width=20  ,bg="yellow"  ,fg="red",font=('times', 15, ' bold ')  )
txt2.place(x=400, y=315)


lbl3 = tk.Label(window, text="Enter Mobile Number",width=20  ,fg="red"  ,bg="yellow"    ,height=2 ,font=('times', 15, ' bold ')) 
lbl3.place(x=100, y=400)

txt3 = tk.Entry(window,width=20  ,bg="yellow"  ,fg="red",font=('times', 15, ' bold ')  )
txt3.place(x=400, y=415)


lbl4 = tk.Label(window, text="Notification : ",width=20  ,fg="red"  ,bg="yellow"  ,height=2 ,font=('times', 15, ' bold underline ')) 
lbl4.place(x=100, y=500)

message = tk.Label(window, text="" ,bg="yellow"  ,fg="red"  ,width=30  ,height=2, activebackground = "yellow" ,font=('times', 15, ' bold ')) 
message.place(x=400, y=500)

def getProfile(Id):
	conn=sqlite3.connect("facebase")
	cmd="SELECT * FROM student WHERE ID="+str(Id)
	cursor=conn.execute(cmd)
	profile=None
	for row in cursor:
		profile=row
	conn.close
	return profile


def getDateColumn():
	wb = load_workbook(filename = "reports.xlsx")
	sheet = wb.get_sheet_by_name('Cse15')
	row_count = sheet.max_row
	column_count = sheet.max_column
	print(row_count)
	print(column_count)
	for i in range(1, column_count + 1):
		col = i
		print(col)
		if sheet.cell(row=1,column=i).value == currentDate:
			#col=col+1
			print(col)
	return col

def createworkbook():
	#create a workbook and add a worksheet
	conn = sqlite3.connect('facebase')
	c = conn.cursor()
	
	if(os.path.exists('./reports.xlsx')):
	    wb = load_workbook(filename = "reports.xlsx")
	    sheet = wb.get_sheet_by_name('Cse15')
	    # sheet[ord() + '1']
	    for col_index in range(1, 100):
	    	col = get_column_letter(col_index)
	    	#print(sheet.cell(row=1,column=col_index).value)
	    	if sheet.cell(row=1,column=col_index).value is None:
	    		#print(sheet.cell(row=1,column=col_index-1).value)
	    		# print sheet.cell('%s%s'% (col2, 1)).value
	    		if sheet.cell(row=1,column=col_index-1).value != currentDate:
	    			sheet.cell(row=1,column=col_index).value = currentDate
	    		break
	 
	    #saving the file
	    wb.save(filename = "reports.xlsx")
	    	
	else:
	    wb = Workbook()
	    dest_filename = 'reports.xlsx'
	    c.execute("SELECT * FROM Student ORDER BY ID ASC")
	    
	    #creating worksheet and giving names to column
	    ws1 = wb.active
	    ws1.title = "Cse15"
	    ws1.append(('ID', 'Name','Mobile', currentDate))
	    ws1.append(('', '', '',''))
	
	    #entering students information from database
	    while True:
	        a = c.fetchone()
	        print(a)
	        if a == None:
	            break
	        else:
	            ws1.append((a[0], a[1],a[2]))
	
	    #saving the file
	    wb.save(filename = dest_filename)
	    


def find_faces(image_path):
    recognizer.read('trainer/trainer.yml')
    image = cv2.imread(image_path)
    # Make a copy to prevent us from modifying the original
    color_img = image.copy()
    filename = os.path.basename(image_path)
    # OpenCV works best with gray images
    gray_img = cv2.cvtColor(color_img, cv2.COLOR_BGR2GRAY)
    # Use OpenCV's built-in Haar classifier
    faces=faceCascade.detectMultiScale(gray_img, 1.2,5)
    print('Number of faces found: {faces}'.format(faces=len(faces)))
    createworkbook()
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb.get_sheet_by_name('Cse15')
    col = getDateColumn()
    print("column")
    print(col)
    for (x, y, w, h) in faces:
        #cv2.rectangle(color_img, (x, y), (x+width, y+height), (0, 255, 0), 2)
        cv2.rectangle(color_img,(x,y),(x+w,y+h),(225,0,0),2)
        Id, conf = recognizer.predict(gray_img[y:y+h,x:x+w])
        print(Id)
        print(conf)
       

        if(conf>50):
        	profile=getProfile(Id)
     	
        else:
        	Id=0
	       	profile=getProfile(Id)
	       	cv2.putText(color_img,'0', (x,y+h),font, fontScale, fontColor)
	       	cv2.putText(color_img,'UNKNOWN', (x,y+h+30),font, fontScale, fontColor)
        if(profile!=None):
        		row_count = sheet.max_row
        		column_count = sheet.max_column
        		print(row_count)
        		print(column_count)
        		for i in range(1, row_count + 1):        			
        			if sheet.cell(row=i,column=1).value == profile[0]:
        				print(sheet.cell(row=i,column=1).value)
        				print(profile[0])
        				print(profile[2])
        				msg="Hi " + profile[1] + " is present today class " + currentDate
        				mobile=str(profile[2])
        				print(mobile)
        				#payload = "sender_id=FSTSMS&message="+msg+"&language=english&route=p&numbers="+mobile
        				#headers = {'authorization': "",'Content-Type': "application/x-www-form-urlencoded",'Cache-Control': "no-cache",}
        				#response = requests.request("POST", url, data=payload, headers=headers)
        				#print(response.text)
        				print(i)
        				print(col)
        				sheet.cell(row=i,column=col).value="Present"
        				wb.save('reports.xlsx')
        				
        		cv2.putText(color_img,str(profile[0]), (x,y+h),font, fontScale, fontColor)
        		cv2.putText(color_img,str(profile[1]), (x,y+h+30),font, fontScale, fontColor)
    cv2.imshow(filename, color_img)
    if cv2.waitKey(1)==ord('q'):
    	cv2.waitKey(0) 
    	cv2.destroyAllWindows()
    	
  

def insertOrUpdate(Id,name,mob):	
	conn= sqlite3.connect("facebase")
	cmd="SELECT ID,Name FROM student WHERE ID="+str(Id)
	cursor=conn.execute(cmd)
	isRecordExist=0
	for row in cursor:
		isRecordExist=1
	if(isRecordExist==1):
	        print("update")
	        print(Id)
	        print(name)
	        cmd="UPDATE student SET Name='"+str(name)+"',Mobile='"+mob+"' WHERE ID="+str(Id)
	        message.configure(text= "Updated Successfully")
	else:
		print("insert")
		print(Id)
		print(name)
		cmd="INSERT INTO student Values("+str(Id)+",'"+str(name)+"','"+mob+"')"
		message.configure(text= "Inserted Successfully")
	conn.execute(cmd)
	conn.commit()
	conn.close() 
def clear():
    print("Clear1")
    txt.delete(0, 'end')    
    txt2.delete(0, 'end')
    txt3.delete(0, 'end')   
    res = ""
    message.configure(text= res)

def clear2():
    print("Clear2") 
    
def addStudent():
    print("addStudent")	
    
    Id=txt.get()
    name=txt2.get()
    mob=txt3.get()
    if((Id!="") and (name!="") and (mob!="") and (len(mob)==10)):
    	print(Id)
    	print(name)
    	print(mob)
    	print("if")
    	insertOrUpdate(Id,name,mob)
    	sampleNum=0
    	cam = cv2.VideoCapture(0)
    
    
    	while True:
    		ret, frame = cam.read()
    		if ret == True:
    			gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    			faces = faceCascade.detectMultiScale(
    		            frame,
    		            scaleFactor=1.1,
    		            minNeighbors=5,
    		            minSize=(30, 30),
    		            #flags=cv2.cv.CV_HAAR_SCALE_IMAGE
        	    	)
    			for (x,y,w,h) in faces:
    				cv2.rectangle(frame,(x,y),(x+w,y+h),(255,0,0),2)
    				sampleNum=sampleNum+1
    				print(Id +'.'+ str(sampleNum))
    				cv2.imwrite("dataSet/User."+Id +'.'+ str(sampleNum) + ".jpg", gray[y:y+h,x:x+w])
    				#cv2.imshow('frame',img)
    		# Display the resulting frame
    		cv2.imshow('Video', frame)
    		if cv2.waitKey(1)==ord('q'):
        	        break
    	cam.release()
    	cv2.destroyAllWindows()
    else:
    	print("else")
    	print("ID")
    	print(Id)
    	print(name)
    	print(mob)
    	print(name.isalpha())
    	if Id=="":
    		res = "Enter ID"
    		message.configure(text= res)
    	elif name=="":
	    	res = "Enter Name"
    		message.configure(text= res)
    	elif mob=="":
    		res = "Enter Mobile Number"
    		message.configure(text= res)
    	elif len(mob)!="":
    		res = "Enter 10 digit Mobile Number"
    		message.configure(text= res)
    	
    	#if name.isalpha():
    	#	print(name.isalpha)
    	#	print("in alpha")
    	#	res = "Enter Alphabetical Name"
    	#	message.configure(text= res)

  
def detection():
    print("detection")
    cam = cv2.VideoCapture(0)
    while True:
    	return_value, image = cam.read()
    	gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    	faces = faceCascade.detectMultiScale(image,scaleFactor=1.1,minNeighbors=5,minSize=(30, 30))
    	for (x,y,w,h) in faces:
    		cv2.rectangle(image,(x,y),(x+w,y+h),(255,0,0),2)
    		cv2.imshow('Video', image)
    	cv2.imwrite('image.png', image)
    	if cv2.waitKey(1)==ord('q'):
    		break
    cam.release()
    cv2.destroyAllWindows()
    find_faces('image.png')
       
    
    
    
def train():
    print("Train Images")
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    
    def getImagesAndLabels(path):
        print(path)
        imagePaths=[os.path.join(path,f) for f in os.listdir(path)] 
        
        faceSamples=[]
        
        Ids=[]
        
        for imagePath in imagePaths:
            pilImage=Image.open(imagePath).convert('L')
            imageNp=np.array(pilImage,'uint8')
            Id=int(os.path.split(imagePath)[-1].split(".")[1])
            faces=faceCascade.detectMultiScale(imageNp)
            for (x,y,w,h) in faces:
                faceSamples.append(imageNp[y:y+h,x:x+w])
                Ids.append(Id)
        return faceSamples,Ids
    
    
    faces,Ids = getImagesAndLabels('dataSet')
    recognizer.train(faces, np.array(Ids))
    recognizer.write('trainer/trainer.yml')
    res = "Image Train is Finished"
    message.configure(text= res)
  
clearButton = tk.Button(window, text="Clear", command=clear  ,fg="white"  ,bg="blue"  ,width=20  ,height=2 ,activebackground = "Red" ,font=('times', 15, ' bold '))
clearButton.place(x=950, y=200)
#clearButton2 = tk.Button(window, text="Clear", command=clear2  ,fg="red"  ,bg="blue"  ,width=20  ,height=2, activebackground = "Red" ,font=('times', 15, ' bold '))
#clearButton2.place(x=950, y=300)    
addst = tk.Button(window, text="Add Student", command=addStudent  ,fg="red"  ,bg="yellow"  ,width=20  ,height=3, activebackground = "Red" ,font=('times', 15, ' bold '))
addst.place(x=100, y=600)
trainImg = tk.Button(window, text="Train Images", command=train  ,fg="red"  ,bg="yellow"  ,width=20  ,height=3, activebackground = "Red" ,font=('times', 15, ' bold '))
trainImg.place(x=400, y=600)
detect = tk.Button(window, text="Detection", command=detection  ,fg="red"  ,bg="yellow"  ,width=20  ,height=3, activebackground = "Red" ,font=('times', 15, ' bold '))
detect.place(x=700, y=600)
quitWindow = tk.Button(window, text="Quit", command=window.destroy  ,fg="red"  ,bg="yellow"  ,width=20  ,height=3, activebackground = "Red" ,font=('times', 15, ' bold '))
quitWindow.place(x=1000, y=600)

 
window.mainloop()